#!/usr/bin/env python3
"""
Build an offline Rush Duel English database from Yugipedia.

Outputs:
  - rush_yugipedia_english.sqlite
  - rush_yugipedia_english.json
  - rush_yugipedia_english.xlsx

The JSON is intended for browser userscripts hosted on GitHub raw pages.
The SQLite file is intended for inspection, re-export, and partial resume.
The XLSX file is intended for quick inspection of page IDs, database IDs, and names.
"""

from __future__ import annotations

import argparse
import datetime as _dt
import gzip
import html
import json
import os
import queue
import random
import re
import sqlite3
import threading
import time
import tkinter as tk
from dataclasses import dataclass
from tkinter import filedialog, messagebox, ttk
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen


API_URL = "https://yugipedia.com/api.php"
DEFAULT_OUTPUT_DIR = os.path.abspath(os.getcwd())
DEFAULT_SQLITE_NAME = "rush_yugipedia_english.sqlite"
DEFAULT_JSON_NAME = "rush_yugipedia_english.json"
DEFAULT_XLSX_NAME = "rush_yugipedia_english.xlsx"

# Replace the contact address before doing repeated full builds.
USER_AGENT = (
    "RushYugipediaDatabaseBuilder/1.1 "
    "(offline cache builder; contact: replace-with-your-email@example.com)"
)

REQUEST_INTERVAL_SECONDS = 1.0
REQUEST_ATTEMPTS = 6
FETCH_BATCH_SIZE = 50

CARD_CATEGORIES = [
    "Rush Duel cards",
]

PRODUCT_CATEGORIES = [
    "Rush Duel Booster Packs",
]

PRODUCT_SEED_TITLES = [
    "Booster pack",
    "Deck Mod Pack",
    "Advance Pack",
    "Maximum Ultra Pack",
    "Over Rush Pack (series)",
    "Triple Build Pack",
    "Battle Pack (Rush Duel)",
    "Secret Ace Pack (series)",
    "Special Victory Pack",
]

RUSH_PRODUCT_LINK_HINTS = [
    "Deck Mod Pack",
    "Advance Pack",
    "Maximum Ultra Pack",
    "Over Rush Pack",
    "Triple Build Pack",
    "Battle Pack",
    "Secret Ace Pack",
    "Special Victory Pack",
    "Event Pack",
]

CARD_NUMBER_RE = re.compile(r"\b((?:RD|RD/[A-Z0-9]+|RD/[A-Z0-9-]+)[A-Z0-9/-]*-JP[A-Z]?\d{3})\b", re.I)
SET_CODE_SUFFIX_RE = re.compile(r"-JP[A-Z]?\d{3}$", re.I)
JP_PREFIX_RE = re.compile(r"\b((?:RD|RD/[A-Z0-9]+|RD/[A-Z0-9-]+)[A-Z0-9/-]*-JP[A-Z]?)\b", re.I)
JP_PREFIX_SUFFIX_RE = re.compile(r"-JP[A-Z]?$", re.I)

CARD_FIELD_HINTS = {
    "database_id",
    "card_type",
    "property",
    "attribute",
    "level",
    "rank",
    "link",
    "pendulum_scale",
    "types",
    "requirement",
    "text",
    "lore",
    "description",
    "jp_sets",
    "ja_sets",
    "en_sets",
}


class StopRequested(Exception):
    pass


class YugipediaError(Exception):
    pass


@dataclass
class PageResult:
    title: str
    pageid: int
    revid: int
    wikitext: str


class RateLimiter:
    """Small process-local rate limiter for polite API access."""

    def __init__(self, seconds_between_requests: float) -> None:
        self.seconds_between_requests = seconds_between_requests
        self.lock = threading.Lock()
        self.next_allowed = 0.0

    def wait(self, stop_event: threading.Event) -> None:
        with self.lock:
            now = time.monotonic()
            delay = self.next_allowed - now
            if delay > 0:
                sleep_with_stop(delay, stop_event)
            self.next_allowed = time.monotonic() + self.seconds_between_requests


RATE_LIMITER = RateLimiter(REQUEST_INTERVAL_SECONDS)


def utc_now() -> str:
    return _dt.datetime.now(_dt.timezone.utc).replace(microsecond=0).isoformat()


def wiki_url(title: str) -> str:
    return f"https://yugipedia.com/wiki/{quote(title.replace(' ', '_'))}"


def sleep_with_stop(seconds: float, stop_event: threading.Event) -> None:
    end = time.monotonic() + seconds
    while True:
        if stop_event.is_set():
            raise StopRequested()
        remaining = end - time.monotonic()
        if remaining <= 0:
            return
        time.sleep(min(0.2, remaining))


def get_retry_after_seconds(exc: HTTPError) -> Optional[float]:
    value = exc.headers.get("Retry-After") if exc.headers else None
    if not value:
        return None
    try:
        return max(0.0, float(value))
    except ValueError:
        return None


def decode_response(response: Any) -> str:
    raw = response.read()
    encoding = (response.headers.get("Content-Encoding") or "").lower()
    if encoding == "gzip":
        raw = gzip.decompress(raw)
    charset = response.headers.get_content_charset() or "utf-8"
    return raw.decode(charset)


def build_request(query: Dict[str, Any]) -> Request:
    encoded = urlencode({key: str(value) for key, value in query.items()}).encode("utf-8")
    headers = {
        "Accept": "application/json",
        "Accept-Encoding": "gzip",
        "Content-Type": "application/x-www-form-urlencoded; charset=utf-8",
        "User-Agent": USER_AGENT,
    }
    if len(encoded) > 1800:
        return Request(API_URL, data=encoded, headers=headers, method="POST")
    return Request(f"{API_URL}?{encoded.decode('utf-8')}", headers=headers, method="GET")


def request_json(query: Dict[str, Any], stop_event: threading.Event, attempts: int = REQUEST_ATTEMPTS) -> Dict[str, Any]:
    query = dict(query)
    query.setdefault("format", "json")
    query.setdefault("formatversion", "2")
    query.setdefault("utf8", "1")
    query.setdefault("maxlag", "5")

    last_error: Optional[BaseException] = None
    for attempt in range(attempts):
        if stop_event.is_set():
            raise StopRequested()
        try:
            RATE_LIMITER.wait(stop_event)
            req = build_request(query)
            with urlopen(req, timeout=45) as response:
                data = json.loads(decode_response(response))

            if isinstance(data, dict) and data.get("error"):
                error = data["error"]
                code = str(error.get("code", "api_error"))
                info = str(error.get("info", "Yugipedia API error"))
                last_error = YugipediaError(f"{code}: {info}")
                if code == "maxlag" and attempt + 1 < attempts:
                    lag = error.get("lag")
                    try:
                        delay = max(5.0, float(lag) + random.random())
                    except (TypeError, ValueError):
                        delay = 7.0 + random.random()
                    sleep_with_stop(delay, stop_event)
                    continue
                raise last_error

            return data
        except StopRequested:
            raise
        except HTTPError as exc:
            last_error = exc
            if attempt + 1 >= attempts:
                break
            retry_after = get_retry_after_seconds(exc)
            delay = retry_after if retry_after is not None else ([1.5, 3.0, 7.0, 12.0, 20.0][min(attempt, 4)] + random.random())
            sleep_with_stop(delay, stop_event)
        except (URLError, TimeoutError, json.JSONDecodeError, OSError) as exc:
            last_error = exc
            if attempt + 1 >= attempts:
                break
            delay = [1.5, 3.0, 7.0, 12.0, 20.0][min(attempt, 4)] + random.random()
            sleep_with_stop(delay, stop_event)
        except Exception as exc:
            last_error = exc
            if attempt + 1 >= attempts:
                break
            delay = [1.5, 3.0, 7.0, 12.0, 20.0][min(attempt, 4)] + random.random()
            sleep_with_stop(delay, stop_event)

    raise YugipediaError(str(last_error))


def category_members(category: str, stop_event: threading.Event, log: Callable[[str], None]) -> List[str]:
    titles: List[str] = []
    cont: Dict[str, str] = {}
    while True:
        query = {
            "action": "query",
            "list": "categorymembers",
            "cmtitle": f"Category:{category}",
            "cmnamespace": "0",
            "cmprop": "ids|title",
            "cmlimit": "500",
        }
        query.update(cont)
        data = request_json(query, stop_event)
        members = data.get("query", {}).get("categorymembers", [])
        titles.extend(member["title"] for member in members if member.get("title"))
        cont = data.get("continue") or {}
        if not cont:
            break
    log(f"Category '{category}': {len(titles)} pages")
    return titles


def search_titles(search: str, stop_event: threading.Event) -> List[str]:
    titles: List[str] = []
    cont: Dict[str, str] = {}
    while True:
        query = {
            "action": "query",
            "list": "search",
            "srnamespace": "0",
            "srlimit": "500",
            "srsearch": search,
        }
        query.update(cont)
        data = request_json(query, stop_event)
        titles.extend(row["title"] for row in data.get("query", {}).get("search", []) if row.get("title"))
        cont = data.get("continue") or {}
        if not cont:
            break
    return titles


def iter_batches(items: Sequence[Any], size: int) -> Iterable[List[Any]]:
    for index in range(0, len(items), size):
        yield list(items[index:index + size])


def normalize_pages(pages: Any) -> Iterable[Dict[str, Any]]:
    if isinstance(pages, dict):
        return pages.values()
    if isinstance(pages, list):
        return pages
    return []


def fetch_pages(titles: Sequence[str], stop_event: threading.Event) -> Dict[str, PageResult]:
    results: Dict[str, PageResult] = {}
    for batch in iter_batches(list(titles), FETCH_BATCH_SIZE):
        if not batch:
            continue
        data = request_json({
            "action": "query",
            "prop": "revisions",
            "rvprop": "ids|content",
            "rvslots": "main",
            "titles": "|".join(batch),
            "redirects": "1",
        }, stop_event)
        pages = data.get("query", {}).get("pages", {})
        redirects = data.get("query", {}).get("redirects", []) or []
        redirect_map = {row.get("to"): row.get("from") for row in redirects if row.get("from") and row.get("to")}

        for page in normalize_pages(pages):
            title = page.get("title") or ""
            revisions = page.get("revisions") or []
            revision = revisions[0] if revisions else {}
            wikitext = get_revision_text(revision)
            if title and wikitext:
                result = PageResult(
                    title=title,
                    pageid=int(page.get("pageid") or 0),
                    revid=int(revision.get("revid") or revision.get("parentid") or 0),
                    wikitext=wikitext,
                )
                results[title] = result
                original_title = redirect_map.get(title)
                if original_title:
                    results[original_title] = result
    return results


def get_revision_text(revision: Dict[str, Any]) -> str:
    slots = revision.get("slots") or {}
    main_slot = slots.get("main") or {}
    return (
        revision.get("*")
        or revision.get("content")
        or main_slot.get("*")
        or main_slot.get("content")
        or ""
    )


def parse_template_fields(wikitext: str) -> Dict[str, str]:
    fields = parse_template_fields_with_mwparser(wikitext)
    if fields:
        return fields
    return parse_template_fields_fallback(wikitext)


def parse_template_fields_with_mwparser(wikitext: str) -> Dict[str, str]:
    try:
        import mwparserfromhell  # type: ignore
    except Exception:
        return {}

    try:
        code = mwparserfromhell.parse(wikitext)
        best_fields: Dict[str, str] = {}
        best_score = -1
        for template in code.filter_templates(recursive=False):
            current: Dict[str, str] = {}
            for param in template.params:
                name = str(param.name).strip()
                if name:
                    current[name] = str(param.value).strip()
            if not current:
                continue
            field_names = {key.lower() for key in current}
            score = len(field_names & CARD_FIELD_HINTS) * 10 + len(current)
            if score > best_score:
                best_fields = current
                best_score = score
        return best_fields
    except Exception:
        return {}


def parse_template_fields_fallback(wikitext: str) -> Dict[str, str]:
    block = extract_best_template_block(wikitext)
    if not block:
        return parse_line_based_fields(wikitext)
    return parse_template_params(block)


def extract_best_template_block(wikitext: str) -> str:
    best = ""
    best_score = -1
    index = 0
    while True:
        start = wikitext.find("{{", index)
        if start < 0:
            break
        end = find_matching_template_end(wikitext, start)
        if end < 0:
            break
        block = wikitext[start + 2:end - 2]
        score = block.count("\n|") * 3 + sum(1 for hint in CARD_FIELD_HINTS if f"|{hint}" in block.lower()) * 10
        if score > best_score:
            best = block
            best_score = score
        index = end
    return best


def find_matching_template_end(text: str, start: int) -> int:
    depth = 0
    index = start
    while index < len(text) - 1:
        pair = text[index:index + 2]
        if pair == "{{":
            depth += 1
            index += 2
            continue
        if pair == "}}":
            depth -= 1
            index += 2
            if depth == 0:
                return index
            continue
        index += 1
    return -1


def split_top_level_pipes(text: str) -> List[str]:
    parts: List[str] = []
    start = 0
    brace_depth = 0
    link_depth = 0
    index = 0
    while index < len(text):
        pair = text[index:index + 2]
        if pair == "{{":
            brace_depth += 1
            index += 2
            continue
        if pair == "}}" and brace_depth:
            brace_depth -= 1
            index += 2
            continue
        if pair == "[[":
            link_depth += 1
            index += 2
            continue
        if pair == "]]" and link_depth:
            link_depth -= 1
            index += 2
            continue
        if text[index] == "|" and brace_depth == 0 and link_depth == 0:
            parts.append(text[start:index])
            start = index + 1
        index += 1
    parts.append(text[start:])
    return parts


def parse_template_params(template_inner: str) -> Dict[str, str]:
    parts = split_top_level_pipes(template_inner)
    fields: Dict[str, str] = {}
    for raw_part in parts[1:]:
        part = raw_part.strip()
        if not part or "=" not in part:
            continue
        key, value = part.split("=", 1)
        key = key.strip()
        if not key or key.isdigit():
            continue
        fields[key] = value.strip()
    return fields


def parse_line_based_fields(wikitext: str) -> Dict[str, str]:
    fields: Dict[str, str] = {}
    current_key: Optional[str] = None
    for line in wikitext.splitlines():
        match = re.match(r"^\|\s*([^=]+?)\s*=\s*(.*)$", line)
        if match:
            current_key = match.group(1).strip()
            fields[current_key] = match.group(2).strip()
        elif current_key and not line.startswith("}}"):
            fields[current_key] += "\n" + line.strip()
        else:
            current_key = None
    return fields


def clean_wiki_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"<!--[\s\S]*?-->", "", text)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"<ref\b[^>]*>[\s\S]*?</ref>", "", text, flags=re.I)
    text = re.sub(r"<ref\b[^/]*/>", "", text, flags=re.I)
    text = text.replace("{{!}}", "|")
    text = replace_wiki_links(text)
    text = reduce_simple_templates(text)
    text = re.sub(r"'''?", "", text)
    text = text.replace("&nbsp;", " ")
    text = html.unescape(text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def replace_wiki_links(text: str) -> str:
    def repl(match: re.Match[str]) -> str:
        page = match.group(1)
        label = match.group(2)
        return label or page

    return re.sub(r"\[\[([^\]|#]+)(?:#[^\]|]*)?(?:\|([^\]]+))?]]", repl, text)


def reduce_simple_templates(text: str) -> str:
    previous = None
    while previous != text:
        previous = text
        text = re.sub(r"\{\{([^{}]*)\}\}", reduce_one_template, text)
    return text


def reduce_one_template(match: re.Match[str]) -> str:
    content = match.group(1)
    parts = [part.strip() for part in split_top_level_pipes(content)]
    if not parts:
        return ""
    name = parts[0].strip().lower()
    params = parts[1:]
    if name in {"ruby", "sic", "Sic".lower()} and params:
        return params[0]
    if name in {"!"}:
        return "|"
    if name in {"nbsp"}:
        return " "
    if name in {"efn", "ref", "citation needed"}:
        return ""
    for param in params:
        if param and "=" not in param:
            return param
    return ""


def get_display_data(title: str, fields: Dict[str, str]) -> Dict[str, Any]:
    card_type = clean_wiki_text(fields.get("card_type", ""))
    property_value = clean_wiki_text(fields.get("property", ""))
    attribute = clean_wiki_text(fields.get("attribute", ""))
    level = clean_wiki_text(fields.get("level", ""))
    rank = clean_wiki_text(fields.get("rank", ""))
    link = clean_wiki_text(fields.get("link", ""))
    pendulum_scale = clean_wiki_text(fields.get("pendulum_scale", ""))
    types = clean_wiki_text(fields.get("types", ""))
    requirement = clean_wiki_text(fields.get("requirement", ""))
    effect = clean_wiki_text(fields.get("text") or fields.get("lore") or fields.get("description") or "")
    preface = [
        clean_wiki_text(fields.get("summon_condition", "")),
        clean_wiki_text(fields.get("summoning_condition", "")),
        clean_wiki_text(fields.get("condition", "")),
        clean_wiki_text(fields.get("materials", "")),
        clean_wiki_text(fields.get("fusion_materials", "")),
    ]
    preface = [item for item in preface if item]
    return {
        "title": clean_wiki_text(title),
        "cardType": card_type,
        "property": property_value,
        "attribute": attribute,
        "level": level,
        "rank": rank,
        "link": link,
        "pendulumScale": pendulum_scale,
        "types": types,
        "requirement": requirement,
        "effect": effect,
        "preface": preface,
        "kindText": get_kind_text(card_type, property_value, types),
    }


def get_kind_text(card_type: str, property_value: str, types: str) -> str:
    if types:
        return " / ".join(part.strip() for part in types.split("/") if part.strip())
    if card_type:
        return f"{property_value} {card_type}".strip()
    return ""


def extract_prints(fields: Dict[str, str], wikitext: str) -> List[Dict[str, str]]:
    set_text = fields.get("jp_sets") or fields.get("ja_sets") or fields.get("en_sets") or ""
    prints: List[Dict[str, str]] = []
    seen: set[str] = set()

    for line in str(set_text).splitlines():
        parts = [clean_wiki_text(part) for part in line.split(";")]
        if not parts or not parts[0]:
            continue
        number = parts[0]
        set_name = parts[1] if len(parts) > 1 else ""
        rarity = parts[2] if len(parts) > 2 else ""
        add_print(prints, seen, number, set_name, rarity, line)

    for match in CARD_NUMBER_RE.finditer(wikitext):
        number = match.group(1)
        start = max(0, match.start() - 180)
        end = min(len(wikitext), match.end() + 260)
        context = wikitext[start:end]
        guessed_set = guess_set_name_from_context(context, number)
        add_print(prints, seen, number, guessed_set, "", clean_wiki_text(context))

    return prints


def add_print(prints: List[Dict[str, str]], seen: set[str], number: str, set_name: str, rarity: str, context: str) -> None:
    clean_number = clean_wiki_text(number).upper()
    if not clean_number or clean_number in seen:
        return
    seen.add(clean_number)
    set_code = SET_CODE_SUFFIX_RE.sub("", clean_number)
    prints.append({
        "number": clean_number,
        "setCode": set_code,
        "setName": clean_wiki_text(set_name),
        "rarity": clean_wiki_text(rarity),
        "rawContext": context,
    })


def guess_set_name_from_context(context: str, number: str) -> str:
    cleaned = replace_wiki_links(context)
    cleaned = cleaned.replace(number, " ")
    chunks = re.split(r"[;\n|=]", cleaned)
    candidates = [clean_wiki_text(chunk) for chunk in chunks]
    candidates = [
        item for item in candidates
        if item and len(item) > 5 and not CARD_NUMBER_RE.search(item) and not item.lower().startswith(("rarity", "set"))
    ]
    return candidates[0] if candidates else ""


def normalize_set_code_from_prefix(prefix: str) -> str:
    return JP_PREFIX_SUFFIX_RE.sub("", clean_wiki_text(prefix).upper())


def extract_jp_prefixes(fields: Dict[str, str], wikitext: str) -> List[str]:
    candidates: List[str] = []
    for key, value in fields.items():
        if key.lower().strip() in {"prefix", "jp_prefix", "ja_prefix"}:
            candidates.extend(match.group(1).upper() for match in JP_PREFIX_RE.finditer(str(value)))
    for match in JP_PREFIX_RE.finditer(wikitext):
        value = match.group(1).upper()
        if not CARD_NUMBER_RE.fullmatch(value):
            candidates.append(value)
    seen: Dict[str, None] = {}
    for value in candidates:
        if value.endswith("-JP") or re.search(r"-JP[A-Z]$", value):
            seen.setdefault(value, None)
    return sorted(seen)


def extract_wiki_link_titles(wikitext: str) -> List[str]:
    titles: List[str] = []
    seen: Dict[str, None] = {}
    for match in re.finditer(r"\[\[([^\]|#]+)(?:#[^\]|]*)?(?:\|[^\]]+)?]]", wikitext):
        title = clean_wiki_text(match.group(1))
        if title and not is_obviously_non_card_title(title) and ":" not in title.split(":", 1)[0]:
            if title not in seen:
                seen[title] = None
                titles.append(title)
    return titles


def is_likely_rush_product_title(title: str) -> bool:
    lowered = title.lower()
    if "(rush duel)" in lowered:
        return True
    return any(hint.lower() in lowered for hint in RUSH_PRODUCT_LINK_HINTS)


def infer_product_type(name: str, page_title: str = "", product_type: str = "") -> str:
    existing = clean_wiki_text(product_type)
    if existing:
        return existing
    text = f"{name} {page_title}".lower()
    if "deck modification pack" in text:
        return "Deck Modification Pack"
    if "booster pack" in text:
        return "Booster Pack"
    if "starter deck" in text:
        return "Starter Deck"
    if "start deck" in text:
        return "Start Deck"
    if "starter pack" in text:
        return "Starter Pack"
    if "start pack" in text:
        return "Start Pack"
    if "structure deck" in text:
        return "Structure Deck"
    if "tournament pack" in text:
        return "Tournament Pack"
    if "promotion" in text or "promotional" in text or "promo" in text:
        return "Promotional"
    if "pack" in text:
        return "Pack"
    if "deck" in text:
        return "Deck"
    return ""


def looks_like_card_page(fields: Dict[str, str]) -> bool:
    normalized = {key.lower() for key in fields}
    return bool(normalized & {"database_id", "card_type", "property", "attribute", "types", "requirement", "lore", "text"})


def get_card_identifiers(page: PageResult, fields: Dict[str, str]) -> Tuple[str, str, str]:
    official_database_id = clean_wiki_text(fields.get("database_id", ""))
    if official_database_id:
        return official_database_id, official_database_id, "official"
    if page.pageid:
        return f"pageid:{page.pageid}", "", "provisional"
    return f"title:{page.title.casefold()}", "", "provisional"


class Database:
    def __init__(self, path: str) -> None:
        self.path = path
        parent = os.path.dirname(path)
        if parent:
            os.makedirs(parent, exist_ok=True)
        self.conn = sqlite3.connect(path, check_same_thread=False)
        self.lock = threading.Lock()
        self.init_schema()

    def init_schema(self) -> None:
        with self.conn:
            self.conn.execute("PRAGMA foreign_keys = ON")
            self.conn.executescript("""
                PRAGMA journal_mode=WAL;
                CREATE TABLE IF NOT EXISTS meta (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS cards (
                    database_id TEXT PRIMARY KEY,
                    official_database_id TEXT,
                    source_status TEXT NOT NULL DEFAULT 'official',
                    title TEXT NOT NULL,
                    page_title TEXT NOT NULL,
                    url TEXT NOT NULL,
                    pageid INTEGER,
                    revid INTEGER,
                    card_type TEXT,
                    property TEXT,
                    attribute TEXT,
                    level TEXT,
                    rank TEXT,
                    link TEXT,
                    pendulum_scale TEXT,
                    types TEXT,
                    kind_text TEXT,
                    requirement TEXT,
                    effect TEXT,
                    preface_json TEXT,
                    fields_json TEXT NOT NULL,
                    display_json TEXT NOT NULL,
                    fetched_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS sets (
                    set_code TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    page_title TEXT,
                    url TEXT,
                    product_type TEXT,
                    release_date TEXT,
                    prefixes_json TEXT,
                    raw_json TEXT,
                    fetched_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS card_prints (
                    database_id TEXT NOT NULL,
                    card_number TEXT NOT NULL,
                    set_code TEXT NOT NULL,
                    set_name TEXT,
                    rarity TEXT,
                    raw_context TEXT,
                    PRIMARY KEY (database_id, card_number),
                    FOREIGN KEY (database_id) REFERENCES cards(database_id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS failures (
                    title TEXT PRIMARY KEY,
                    kind TEXT NOT NULL,
                    error TEXT NOT NULL,
                    last_attempt_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS skipped_pages (
                    title TEXT PRIMARY KEY,
                    kind TEXT NOT NULL,
                    reason TEXT NOT NULL,
                    fetched_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_cards_title ON cards(title);
                CREATE INDEX IF NOT EXISTS idx_cards_page_title ON cards(page_title);
                CREATE INDEX IF NOT EXISTS idx_cards_pageid ON cards(pageid);
                CREATE INDEX IF NOT EXISTS idx_prints_set_code ON card_prints(set_code);
            """)
            self.migrate_schema()
            self.conn.execute("INSERT OR REPLACE INTO meta(key, value) VALUES (?, ?)", ("schema_version", "3"))

    def migrate_schema(self) -> None:
        card_columns = {row[1] for row in self.conn.execute("PRAGMA table_info(cards)")}
        if "official_database_id" not in card_columns:
            self.conn.execute("ALTER TABLE cards ADD COLUMN official_database_id TEXT")
            self.conn.execute("""
                UPDATE cards
                SET official_database_id = database_id
                WHERE database_id NOT LIKE 'pageid:%' AND database_id NOT LIKE 'title:%'
            """)
        if "source_status" not in card_columns:
            self.conn.execute("ALTER TABLE cards ADD COLUMN source_status TEXT NOT NULL DEFAULT 'official'")
            self.conn.execute("""
                UPDATE cards
                SET source_status = CASE
                    WHEN database_id LIKE 'pageid:%' OR database_id LIKE 'title:%' THEN 'provisional'
                    ELSE 'official'
                END
            """)
        set_columns = {row[1] for row in self.conn.execute("PRAGMA table_info(sets)")}
        if "prefixes_json" not in set_columns:
            self.conn.execute("ALTER TABLE sets ADD COLUMN prefixes_json TEXT")

    def is_card_revision_current(self, page_title: str, revid: int) -> bool:
        if not revid:
            return False
        with self.lock:
            row = self.conn.execute(
                "SELECT revid FROM cards WHERE page_title = ?",
                (page_title,),
            ).fetchone()
        return bool(row and int(row[0] or 0) == revid)

    def upsert_card(self, page: PageResult, fields: Dict[str, str], display: Dict[str, Any], prints: List[Dict[str, str]]) -> str:
        database_id, official_database_id, source_status = get_card_identifiers(page, fields)
        fetched_at = utc_now()
        with self.lock, self.conn:
            self.conn.execute(
                "DELETE FROM cards WHERE page_title = ? AND database_id <> ?",
                (page.title, database_id),
            )
            self.conn.execute("""
                INSERT OR REPLACE INTO cards (
                    database_id, official_database_id, source_status, title, page_title, url, pageid, revid, card_type, property,
                    attribute, level, rank, link, pendulum_scale, types, kind_text, requirement,
                    effect, preface_json, fields_json, display_json, fetched_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                database_id, official_database_id, source_status,
                display["title"], page.title, wiki_url(page.title), page.pageid, page.revid,
                display["cardType"], display["property"], display["attribute"], display["level"],
                display["rank"], display["link"], display["pendulumScale"], display["types"],
                display["kindText"], display["requirement"], display["effect"],
                json.dumps(display["preface"], ensure_ascii=False),
                json.dumps(fields, ensure_ascii=False, sort_keys=True),
                json.dumps(display, ensure_ascii=False, sort_keys=True),
                fetched_at,
            ))
            self.conn.execute("DELETE FROM card_prints WHERE database_id = ?", (database_id,))
            for item in prints:
                self.conn.execute("""
                    INSERT OR REPLACE INTO card_prints(database_id, card_number, set_code, set_name, rarity, raw_context)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    database_id, item["number"], item["setCode"], item["setName"], item["rarity"], item["rawContext"],
                ))
                if item["setCode"] and item["setName"]:
                    self.conn.execute("""
                        INSERT OR IGNORE INTO sets(set_code, name, page_title, url, product_type, release_date, prefixes_json, raw_json, fetched_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        item["setCode"], apply_set_suffixes(item["setName"], item["rawContext"]), "",
                        "", "", "", "", json.dumps(item, ensure_ascii=False), fetched_at,
                    ))
            self.conn.execute("DELETE FROM failures WHERE title = ?", (page.title,))
            self.conn.execute("DELETE FROM skipped_pages WHERE title = ?", (page.title,))
        return database_id

    def upsert_product_page(self, page: PageResult, fields: Dict[str, str]) -> int:
        text = page.wikitext
        numbers = sorted({match.group(1).upper() for match in CARD_NUMBER_RE.finditer(text)})
        prefixes = extract_jp_prefixes(fields, text)
        set_codes = sorted({SET_CODE_SUFFIX_RE.sub("", number) for number in numbers} | {normalize_set_code_from_prefix(prefix) for prefix in prefixes})
        fetched_at = utc_now()
        with self.lock, self.conn:
            for code in set_codes:
                code_prefixes = [prefix for prefix in prefixes if normalize_set_code_from_prefix(prefix) == code]
                name = clean_wiki_text(page.title)
                self.conn.execute("""
                    INSERT OR REPLACE INTO sets(set_code, name, page_title, url, product_type, release_date, prefixes_json, raw_json, fetched_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    code, name, page.title, wiki_url(page.title), clean_wiki_text(fields.get("type", "")),
                    clean_wiki_text(fields.get("date", "") or fields.get("release_date", "")),
                    json.dumps(code_prefixes, ensure_ascii=False),
                    json.dumps(fields, ensure_ascii=False, sort_keys=True), fetched_at,
                ))
            self.conn.execute("DELETE FROM failures WHERE title = ?", (page.title,))
            self.conn.execute("DELETE FROM skipped_pages WHERE title = ?", (page.title,))
        return len(set_codes)

    def record_failure(self, title: str, kind: str, error: BaseException) -> None:
        with self.lock, self.conn:
            self.conn.execute("""
                INSERT OR REPLACE INTO failures(title, kind, error, last_attempt_at)
                VALUES (?, ?, ?, ?)
            """, (title, kind, str(error), utc_now()))

    def record_skip(self, title: str, kind: str, reason: str) -> None:
        with self.lock, self.conn:
            self.conn.execute("""
                INSERT OR REPLACE INTO skipped_pages(title, kind, reason, fetched_at)
                VALUES (?, ?, ?, ?)
            """, (title, kind, reason, utc_now()))

    def failed_titles(self) -> List[Tuple[str, str]]:
        with self.lock:
            return list(self.conn.execute("SELECT title, kind FROM failures ORDER BY last_attempt_at DESC"))

    def export_json(self, path: str) -> Tuple[int, int]:
        parent = os.path.dirname(path)
        if parent:
            os.makedirs(parent, exist_ok=True)
        with self.lock:
            cards = list(self.conn.execute("""
                SELECT database_id, official_database_id, source_status, title, url, pageid, page_title, fields_json, display_json
                FROM cards
                ORDER BY title COLLATE NOCASE
            """))
            sets = list(self.conn.execute("""
                SELECT
                    s.set_code,
                    s.name,
                    s.page_title,
                    s.url,
                    s.product_type,
                    s.release_date,
                    s.prefixes_json,
                    COUNT(cp.card_number) AS card_print_count
                FROM sets AS s
                LEFT JOIN card_prints AS cp ON cp.set_code = s.set_code
                GROUP BY s.set_code, s.name, s.page_title, s.url, s.product_type, s.release_date, s.prefixes_json
                ORDER BY s.set_code
            """))
            prints = list(self.conn.execute("SELECT database_id, card_number, set_code, set_name, rarity FROM card_prints ORDER BY database_id, card_number"))

        cards_by_id: Dict[str, Dict[str, Any]] = {}
        for database_id, official_database_id, source_status, title, url, pageid, page_title, fields_json, display_json in cards:
            cards_by_id[str(database_id)] = {
                "title": title,
                "url": url,
                "pageid": int(pageid or 0),
                "pageTitle": page_title,
                "officialDatabaseId": official_database_id or "",
                "sourceStatus": source_status or "official",
                "fields": json.loads(fields_json),
                "display": json.loads(display_json),
            }

        sets_by_code: Dict[str, Dict[str, Any]] = {}
        for code, name, page_title, url, product_type, release_date, prefixes_json, card_print_count in sets:
            prefixes = json.loads(prefixes_json or "[]")
            sets_by_code[code] = {
                "name": name,
                "pageTitle": page_title or "",
                "url": url or (wiki_url(page_title) if page_title else ""),
                "productType": infer_product_type(name or "", page_title or "", product_type or ""),
                "releaseDate": release_date or "",
                "prefixes": prefixes,
                "cardPrintCount": int(card_print_count or 0),
            }

        prints_by_card: Dict[str, List[Dict[str, str]]] = {}
        for database_id, card_number, set_code, set_name, rarity in prints:
            prints_by_card.setdefault(str(database_id), []).append({
                "number": card_number,
                "setCode": set_code,
                "setName": set_name or sets_by_code.get(set_code, {}).get("name", ""),
                "rarity": rarity or "",
            })

        data = {
            "schemaVersion": 3,
            "generatedAt": utc_now(),
            "source": "Yugipedia",
            "cardsByDatabaseId": cards_by_id,
            "setsByCode": sets_by_code,
            "printsByDatabaseId": prints_by_card,
        }
        tmp_path = path + ".tmp"
        with open(tmp_path, "w", encoding="utf-8", newline="\n") as handle:
            json.dump(data, handle, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        os.replace(tmp_path, path)
        return len(cards_by_id), len(sets_by_code)

    def export_xlsx(self, path: str) -> int:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ImportError as exc:
            raise YugipediaError("openpyxl is required for XLSX export. Install it with: pip install openpyxl") from exc

        parent = os.path.dirname(path)
        if parent:
            os.makedirs(parent, exist_ok=True)

        with self.lock:
            rows = list(self.conn.execute("""
                SELECT
                    pageid, database_id, official_database_id, source_status, title, page_title, url,
                    card_type, property, attribute, level, types,
                    kind_text, requirement, effect, preface_json
                FROM cards
                ORDER BY title COLLATE NOCASE
            """))
            failures = list(self.conn.execute("""
                SELECT title, kind, error, last_attempt_at
                FROM failures
                ORDER BY last_attempt_at DESC, title COLLATE NOCASE
            """))
            products = list(self.conn.execute("""
                SELECT
                    s.set_code,
                    s.name,
                    s.page_title,
                    s.url,
                    s.product_type,
                    s.release_date,
                    s.prefixes_json,
                    COUNT(cp.card_number) AS card_print_count
                FROM sets AS s
                LEFT JOIN card_prints AS cp ON cp.set_code = s.set_code
                GROUP BY s.set_code, s.name, s.page_title, s.url, s.product_type, s.release_date, s.prefixes_json
                ORDER BY s.set_code
            """))

        wb = Workbook()
        ws = wb.active
        ws.title = "Rush Duel Cards"
        headers = [
            "pageid",
            "database_id",
            "official_database_id",
            "source_status",
            "name",
            "page_title",
            "url",
            "card_type",
            "property",
            "attribute",
            "level",
            "types",
            "kind_text",
            "requirement",
            "effect",
            "preface",
        ]
        ws.append(headers)
        for row in rows:
            values = list(row)
            if not values[7] and values[11]:
                values[7] = "Monster"
            try:
                preface_items = json.loads(values[-1] or "[]")
                values[-1] = "\n".join(str(item) for item in preface_items if item)
            except (TypeError, ValueError, json.JSONDecodeError):
                values[-1] = values[-1] or ""
            ws.append(values)

        if rows:
            last_column = get_column_letter(len(headers))
            table_ref = f"A1:{last_column}{len(rows) + 1}"
            table = Table(displayName="RushDuelCards", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

        ws.freeze_panes = "A2"
        widths = {
            "A": 12,
            "B": 14,
            "C": 20,
            "D": 16,
            "E": 42,
            "F": 42,
            "G": 64,
            "H": 18,
            "I": 16,
            "J": 14,
            "K": 10,
            "L": 30,
            "M": 36,
            "N": 56,
            "O": 90,
            "P": 56,
        }
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
        for row in ws.iter_rows(min_row=2, min_col=14, max_col=16):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        products_ws = wb.create_sheet("Sets and Packs")
        product_headers = [
            "set_code",
            "name",
            "jp_prefixes",
            "product_type",
            "page_title",
            "url",
            "release_date",
            "card_print_count",
        ]
        products_ws.append(product_headers)
        for set_code, name, page_title, url, product_type, release_date, prefixes_json, card_print_count in products:
            prefixes = json.loads(prefixes_json or "[]")
            products_ws.append([
                set_code,
                name,
                "\n".join(prefixes),
                infer_product_type(name or "", page_title or "", product_type or ""),
                page_title or "",
                url or (wiki_url(page_title) if page_title else ""),
                release_date or "",
                int(card_print_count or 0),
            ])
        if products:
            table_ref = f"A1:H{len(products) + 1}"
            table = Table(displayName="SetsAndPacks", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            products_ws.add_table(table)
        products_ws.freeze_panes = "A2"
        product_widths = {"A": 18, "B": 58, "C": 24, "D": 24, "E": 44, "F": 72, "G": 18, "H": 18}
        for column, width in product_widths.items():
            products_ws.column_dimensions[column].width = width
        for row in products_ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        failures_ws = wb.create_sheet("Failed Scrapings")
        failure_headers = ["title", "kind", "error", "last_attempt_at", "url"]
        failures_ws.append(failure_headers)
        for title, kind, error, last_attempt_at in failures:
            failures_ws.append([title, kind, error, last_attempt_at, wiki_url(title)])
        if failures:
            table_ref = f"A1:E{len(failures) + 1}"
            table = Table(displayName="FailedScrapings", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium3",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            failures_ws.add_table(table)
        failures_ws.freeze_panes = "A2"
        failure_widths = {"A": 44, "B": 14, "C": 72, "D": 24, "E": 72}
        for column, width in failure_widths.items():
            failures_ws.column_dimensions[column].width = width
        for row in failures_ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        wb.save(path)
        return len(rows)

    def close(self) -> None:
        self.conn.close()


def apply_set_suffixes(english_name: str, raw_context: str) -> str:
    if "特典カード" in raw_context and "Promo Cards" not in english_name:
        return f"{english_name} (Promo Cards)"
    return english_name


class BuilderWorker:
    def __init__(self, app: "App") -> None:
        self.app = app
        self.stop_event = threading.Event()
        self.thread: Optional[threading.Thread] = None

    def start(self, retry_only: bool = False) -> None:
        if self.thread and self.thread.is_alive():
            return
        self.stop_event.clear()
        self.thread = threading.Thread(target=self.run, args=(retry_only,), daemon=True)
        self.thread.start()

    def stop(self) -> None:
        self.stop_event.set()

    def run(self, retry_only: bool) -> None:
        try:
            self.app.worker_started()
            output_dir = self.app.output_dir.get().strip() or DEFAULT_OUTPUT_DIR
            os.makedirs(output_dir, exist_ok=True)
            db_path = os.path.join(output_dir, DEFAULT_SQLITE_NAME)
            json_path = os.path.join(output_dir, DEFAULT_JSON_NAME)
            xlsx_path = os.path.join(output_dir, DEFAULT_XLSX_NAME)
            db = Database(db_path)
            try:
                limit = self.app.get_limit()
                skip_unchanged = self.app.skip_unchanged.get()
                if retry_only:
                    targets = db.failed_titles()
                    self.app.log(f"Retrying {len(targets)} failed pages")
                    self.process_targets(db, targets, limit, skip_unchanged)
                else:
                    card_titles = discover_card_titles(self.stop_event, self.app.log)
                    product_titles = discover_product_titles(self.stop_event, self.app.log)
                    targets = [(title, "card") for title in card_titles] + [(title, "product") for title in product_titles]
                    if limit:
                        targets = targets[:limit]
                    self.app.log(f"Total queued pages: {len(targets)}")
                    self.process_targets(db, targets, None, skip_unchanged)
                card_count, set_count = db.export_json(json_path)
                xlsx_count = db.export_xlsx(xlsx_path)
            finally:
                db.close()
            self.app.log(f"Exported {card_count} cards and {set_count} sets to {json_path}")
            self.app.log(f"Exported {xlsx_count} card rows to {xlsx_path}")
            self.app.worker_done()
        except StopRequested:
            self.app.log("Stopped by user.")
            self.app.worker_done()
        except Exception as exc:
            self.app.log(f"ERROR: {exc}")
            self.app.worker_done()

    def process_targets(self, db: Database, targets: List[Tuple[str, str]], limit: Optional[int], skip_unchanged: bool) -> None:
        if limit:
            targets = targets[:limit]
        total = len(targets)
        done = 0

        for batch in iter_batches(targets, FETCH_BATCH_SIZE):
            if self.stop_event.is_set():
                raise StopRequested()
            titles = [title for title, _kind in batch]
            pages = fetch_pages(titles, self.stop_event)
            for title, kind in batch:
                done += 1
                try:
                    page = pages.get(title)
                    if not page:
                        raise YugipediaError("Page content not found")
                    message = self.process_page(db, page, kind, skip_unchanged)
                    self.app.log(message)
                except Exception as exc:
                    db.record_failure(title, kind, exc)
                    self.app.log(f"FAILED {kind}: {title}: {exc}")
                self.app.progress(done, total, title)

    def process_page(self, db: Database, page: PageResult, kind: str, skip_unchanged: bool) -> str:
        fields = parse_template_fields(page.wikitext)
        if kind == "product":
            count = db.upsert_product_page(page, fields)
            return f"Product: {page.title} ({count} set codes)"

        if not looks_like_card_page(fields):
            db.record_skip(page.title, kind, "no recognizable card template fields")
            return f"Skipped non-card page: {page.title}"

        if skip_unchanged and db.is_card_revision_current(page.title, page.revid):
            db.record_skip(page.title, kind, "unchanged revision")
            return f"Skipped unchanged card: {page.title}"

        display = get_display_data(page.title, fields)
        prints = extract_prints(fields, page.wikitext)
        database_id = db.upsert_card(page, fields, display, prints)
        return f"Card {database_id}: {display['title']}"


def discover_card_titles(stop_event: threading.Event, log: Callable[[str], None]) -> List[str]:
    seen: Dict[str, None] = {}
    for category in CARD_CATEGORIES:
        try:
            for title in category_members(category, stop_event, log):
                if not is_obviously_non_card_title(title):
                    seen.setdefault(title, None)
        except Exception as exc:
            log(f"Category failed '{category}': {exc}")
    if not seen:
        log("Category discovery found no cards; trying search fallback")
        for title in search_titles('incategory:"Rush Duel cards"', stop_event):
            seen.setdefault(title, None)
    return sorted(seen)


def is_obviously_non_card_title(title: str) -> bool:
    return title.startswith("List of ") or title.startswith("Category:") or title.startswith("Template:")


def discover_product_titles(stop_event: threading.Event, log: Callable[[str], None]) -> List[str]:
    seen: Dict[str, None] = {}
    for title in PRODUCT_SEED_TITLES:
        seen.setdefault(title, None)
    for category in PRODUCT_CATEGORIES:
        try:
            for title in category_members(category, stop_event, log):
                seen.setdefault(title, None)
        except Exception as exc:
            log(f"Product category failed '{category}': {exc}")
    try:
        seed_pages = fetch_pages(PRODUCT_SEED_TITLES, stop_event)
        linked_titles: Dict[str, None] = {}
        for page in seed_pages.values():
            for title in extract_wiki_link_titles(page.wikitext):
                if is_likely_rush_product_title(title):
                    linked_titles.setdefault(title, None)
        log(f"Product seed pages: {len(linked_titles)} linked product candidates")
        for title in linked_titles:
            seen.setdefault(title, None)
    except Exception as exc:
        log(f"Product seed page discovery failed: {exc}")
    return sorted(seen)


class App:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Rush Yugipedia Database Builder")
        self.output_dir = tk.StringVar(value=DEFAULT_OUTPUT_DIR)
        self.limit_var = tk.StringVar(value="")
        self.skip_unchanged = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="Idle")
        self.progress_var = tk.DoubleVar(value=0)
        self.messages: "queue.Queue[Tuple[str, Any]]" = queue.Queue()
        self.worker = BuilderWorker(self)
        self.build_ui()
        self.root.after(100, self.drain_messages)

    def build_ui(self) -> None:
        frame = ttk.Frame(self.root, padding=10)
        frame.grid(row=0, column=0, sticky="nsew")
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(6, weight=1)

        ttk.Label(frame, text="Output folder").grid(row=0, column=0, sticky="w")
        ttk.Entry(frame, textvariable=self.output_dir).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(frame, text="Browse", command=self.browse).grid(row=0, column=2)

        ttk.Label(frame, text="Test limit").grid(row=1, column=0, sticky="w")
        ttk.Entry(frame, textvariable=self.limit_var, width=12).grid(row=1, column=1, sticky="w", padx=6)
        ttk.Label(frame, text="blank = full run").grid(row=1, column=1, sticky="w", padx=110)

        ttk.Checkbutton(
            frame,
            text="Skip unchanged card revisions already in SQLite",
            variable=self.skip_unchanged,
        ).grid(row=2, column=0, columnspan=3, sticky="w")

        buttons = ttk.Frame(frame)
        buttons.grid(row=3, column=0, columnspan=3, sticky="ew", pady=8)
        self.start_button = ttk.Button(buttons, text="Start Full Build", command=lambda: self.worker.start(False))
        self.start_button.pack(side="left")
        self.retry_button = ttk.Button(buttons, text="Retry Failed", command=lambda: self.worker.start(True))
        self.retry_button.pack(side="left", padx=6)
        self.stop_button = ttk.Button(buttons, text="Stop", command=self.worker.stop)
        self.stop_button.pack(side="left")
        self.export_button = ttk.Button(buttons, text="Export JSON + XLSX", command=self.export_only)
        self.export_button.pack(side="left", padx=6)

        ttk.Label(frame, textvariable=self.status_var).grid(row=4, column=0, columnspan=3, sticky="w")
        self.progress_bar = ttk.Progressbar(frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(0, 8))

        self.log_box = tk.Text(frame, width=110, height=28, wrap="word")
        self.log_box.grid(row=6, column=0, columnspan=3, sticky="nsew")
        scroll = ttk.Scrollbar(frame, command=self.log_box.yview)
        scroll.grid(row=6, column=3, sticky="ns")
        self.log_box.configure(yscrollcommand=scroll.set)

    def browse(self) -> None:
        selected = filedialog.askdirectory(initialdir=self.output_dir.get() or DEFAULT_OUTPUT_DIR)
        if selected:
            self.output_dir.set(selected)

    def get_limit(self) -> Optional[int]:
        text = self.limit_var.get().strip()
        if not text:
            return None
        value = int(text)
        if value <= 0:
            raise ValueError("Test limit must be a positive integer or blank")
        return value

    def log(self, message: str) -> None:
        self.messages.put(("log", message))

    def progress(self, done: int, total: int, title: str) -> None:
        self.messages.put(("progress", (done, total, title)))

    def worker_started(self) -> None:
        self.messages.put(("started", None))

    def worker_done(self) -> None:
        self.messages.put(("done", None))

    def drain_messages(self) -> None:
        try:
            while True:
                kind, payload = self.messages.get_nowait()
                if kind == "log":
                    self.log_box.insert("end", f"{payload}\n")
                    self.log_box.see("end")
                elif kind == "progress":
                    done, total, title = payload
                    self.progress_var.set((done / total * 100) if total else 0)
                    self.status_var.set(f"{done}/{total}: {title}")
                elif kind == "started":
                    self.start_button.configure(state="disabled")
                    self.retry_button.configure(state="disabled")
                    self.export_button.configure(state="disabled")
                    self.status_var.set("Running")
                elif kind == "done":
                    self.start_button.configure(state="normal")
                    self.retry_button.configure(state="normal")
                    self.export_button.configure(state="normal")
                    self.status_var.set("Idle")
        except queue.Empty:
            pass
        self.root.after(100, self.drain_messages)

    def export_only(self) -> None:
        try:
            output_dir = self.output_dir.get().strip() or DEFAULT_OUTPUT_DIR
            os.makedirs(output_dir, exist_ok=True)
            db = Database(os.path.join(output_dir, DEFAULT_SQLITE_NAME))
            try:
                cards, sets = db.export_json(os.path.join(output_dir, DEFAULT_JSON_NAME))
                xlsx_rows = db.export_xlsx(os.path.join(output_dir, DEFAULT_XLSX_NAME))
            finally:
                db.close()
            self.log(f"Exported {cards} cards and {sets} sets.")
            self.log(f"Exported {xlsx_rows} card rows to XLSX.")
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))


def run_cli(args: argparse.Namespace) -> None:
    stop_event = threading.Event()
    output_dir = args.output_dir or DEFAULT_OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)
    db = Database(os.path.join(output_dir, DEFAULT_SQLITE_NAME))

    def log(message: str) -> None:
        print(message, flush=True)

    try:
        if args.retry_failed:
            targets = db.failed_titles()
            log(f"Retrying {len(targets)} failed pages")
        else:
            card_titles = discover_card_titles(stop_event, log)
            product_titles = discover_product_titles(stop_event, log)
            targets = [(title, "card") for title in card_titles] + [(title, "product") for title in product_titles]
        if args.limit:
            targets = targets[:args.limit]
        log(f"Total queued pages: {len(targets)}")

        total = len(targets)
        done = 0
        for batch in iter_batches(targets, FETCH_BATCH_SIZE):
            pages = fetch_pages([title for title, _kind in batch], stop_event)
            for title, kind in batch:
                done += 1
                try:
                    page = pages.get(title)
                    if not page:
                        raise YugipediaError("Page content not found")
                    fields = parse_template_fields(page.wikitext)
                    if kind == "product":
                        count = db.upsert_product_page(page, fields)
                        log(f"{done}/{total} Product: {title} ({count} set codes)")
                    else:
                        if not looks_like_card_page(fields):
                            db.record_skip(page.title, kind, "no recognizable card template fields")
                            log(f"{done}/{total} Skipped non-card page: {title}")
                            continue
                        if args.skip_unchanged and db.is_card_revision_current(page.title, page.revid):
                            db.record_skip(page.title, kind, "unchanged revision")
                            log(f"{done}/{total} Skipped unchanged card: {title}")
                            continue
                        display = get_display_data(page.title, fields)
                        prints = extract_prints(fields, page.wikitext)
                        database_id = db.upsert_card(page, fields, display, prints)
                        log(f"{done}/{total} Card {database_id}: {display['title']}")
                except Exception as exc:
                    db.record_failure(title, kind, exc)
                    log(f"{done}/{total} FAILED {kind}: {title}: {exc}")

        card_count, set_count = db.export_json(os.path.join(output_dir, DEFAULT_JSON_NAME))
        xlsx_count = db.export_xlsx(os.path.join(output_dir, DEFAULT_XLSX_NAME))
        log(f"Exported {card_count} cards and {set_count} sets")
        log(f"Exported {xlsx_count} card rows to {os.path.join(output_dir, DEFAULT_XLSX_NAME)}")
    finally:
        db.close()


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build an offline Rush Duel database from Yugipedia.")
    parser.add_argument("--cli", action="store_true", help="Run without the Tkinter GUI")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="Directory for SQLite, JSON, and XLSX outputs")
    parser.add_argument("--limit", type=int, default=None, help="Optional test limit")
    parser.add_argument("--retry-failed", action="store_true", help="Retry pages listed in the failures table")
    parser.add_argument("--no-skip-unchanged", dest="skip_unchanged", action="store_false", help="Do not skip card pages whose revision is already stored")
    parser.set_defaults(skip_unchanged=True)
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()
    if args.cli:
        run_cli(args)
        return
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
